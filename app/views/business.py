#-*- coding: utf-8 -*
"""view for page transfer."""

import logging
from datetime import datetime
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.forms.models import model_to_dict
import apps.settings as settings
from django.http import HttpResponse
from app.views.base import json_extract, to_json, dictfetchall
from app.util.json import decimal_default_proc, parse_parameters_asjson, parse_post_parameters_asjson
from django.db import connection
import json
import os
from os.path import basename
import urllib, zipfile, shutil, tempfile
import openpyxl
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar

logger = logging.getLogger(__name__)

@login_required
def create_file(request):
    """page to portal page."""

    return render(request, "app/business/create_file.html")

@login_required
def get_list(request):

    sql = f"""
        select
        a.emp_id
        , a.first_name
        , a.last_name
        , a.first_name_kana
        , a.last_name_kana
        , a.price
        , b.name project_name
        , b.start_date
        , b.end_date
        , c.id customer_id
        , c.name customer_name
        , c.zip customer_zip
        , c.address customer_address
        , c.tel_no customer_tel_no
        , c.email customer_email
        , c.partener 
    from
        employee a 
    inner join emp_project d
        on a.emp_id = d.emp_id
        and d.current_flag = 1
    inner join project b 
        on d.project_id = b.id
    inner join customer c 
        on b.customer_id = c.id 
    """
    datas = []
    with connection.cursor() as cursor:
        cursor.execute(sql)
        datas = dictfetchall(cursor)

    return HttpResponse(to_json(datas), content_type='application/json')

def months_between(start_date, end_date):
    
    if start_date > end_date:
        raise ValueError(f"Start date {start_date} is not before end date {end_date}")
    
    start = datetime.strptime(start_date[0:6], '%Y%m').date()
    end = datetime.strptime(end_date[:6], '%Y%m').date()

    ret = []
    while start <= end:
        ret.append(datetime.strftime(start, '%Y%m'))
        start += relativedelta(months=1)
    return ret

@login_required
def export(request, service="000", start="20200401", end="20301231"):

    months = months_between(start, end)
    download_file_list = []
    zip_tempfile_path = '{}-{}.zip'.format(start, end)

    sql_export = f"""
        select
            a.emp_id
            , a.first_name
            , a.last_name
            , a.first_name_kana
            , a.last_name_kana
            , a.price * 10000 as price
            , b.name project_name
            , b.start_date
            , b.end_date
            , b.min_time
            , b.max_time
            , c.id customer_id
            , c.name customer_name
            , c.zip customer_zip
            , c.address customer_address
            , c.tel_no customer_tel_no
            , c.email customer_email
            , c.partener 
        from
        employee a 
        inner join emp_project d 
            on a.emp_id = d.emp_id 
        and d.current_flag = 1
        inner join project b 
            on b.id = d.project_id 
        inner join customer c 
            on b.customer_id = c.id 
            and c.partener = %s
        where b.id <> 1
        and date_format(b.start_date, '%%Y%%m') >= %s
        and (b.end_date is null or date_format(b.end_date, '%%Y%%m') <= %s)
        order by c.id
    """
    
    seisan_datas = []
    seisan_sql = """
        select emp_id, seisan_time, ym from seisan where ym >= %s and ym <= %s
    """
    
    with connection.cursor() as cursor:
        cursor.execute(seisan_sql, (start[:6], end[:6]))
        seisan_datas = dictfetchall(cursor)

    print("============")
    print(seisan_datas)
    print("============")
    # 請求書
    if service[0:1] == "1":
        bill_file_datas = []
        with connection.cursor() as cursor:
            cursor.execute(sql_export, (1,start[:6], end[:6]))
            bill_file_datas = dictfetchall(cursor)

        for data in bill_file_datas:
            data['start_date'] = data['start_date'].strftime('%Y%m%d')
            if data['end_date'] is not None:
                data['end_date'] = data['end_date'].strftime('%Y%m%d')

        output_max_month = months[-1]
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if len(bill_file_datas) > 0:
            for month in months:
                # load workbook if you want to write in existing file else use openpyxl.Workbook()
                wb = openpyxl.load_workbook(os.path.join(base_path, "請求書.xlsx"))
                #get the active sheet
                sh1=wb.active
                sh1=wb['請求書']
                done_customer_list = []
                for data1 in bill_file_datas:
                    current_customer_id = data1['customer_id']
                    if current_customer_id in done_customer_list:
                        continue
                    done_customer_list.append(current_customer_id)
                    filename = "{}_{}_請求書.xlsx".format(month, data1['customer_name'])
                    count = 0
                    data_exists_flag = False

                    for data2 in bill_file_datas:
                        if current_customer_id == data2['customer_id'] and month >= data2['start_date'][:6] and (data['end_date'] is None or month <= data2['end_date'][:6]):

                            data_exists_flag = True
                            # pass which row and column and value which you want to update
                            sh1.cell(row=2,column=1,value=data2['customer_name'] + "御中")
                            sh1.cell(row=2,column=8,value='JCL-' + month + "-" + str(count + 1).zfill(3))
                            sh1.cell(row=3,column=8,value=month[:4] + "/" + month[4:] + "/01")
                            sh1.cell(row=4,column=1,value="〒" + data2['customer_zip'][:3] + "-" + data2['customer_zip'][3:])
                            sh1.cell(row=5,column=1,value=data2['customer_address'])
                            sh1.cell(row=8,column=1,value="件名：" + data2['project_name'])
                            # 支払期限
                            current_date = datetime.strptime(month + '01', '%Y%m%d')
                            next_two_month = current_date + relativedelta(months=2)
                            next_month = next_two_month.replace(day=28) + timedelta(days=4)
                            last_day_of_month = next_month - timedelta(days=next_month.day)

                            sh1.cell(row=12,column=8,value=datetime.strftime(last_day_of_month, '%Y%m%d'))
                            sh1.cell(row=15+count,column=1,value=count + 1)
                            sh1.cell(row=15+count,column=2,value=data2['project_name'] + "【" + data2['first_name'] + data2['last_name'] +  "】")
                            sh1.cell(row=15+count,column=5,value=1)
                            sh1.cell(row=15+count,column=6,value="人・月")
                            sh1.cell(row=15+count,column=7,value=data2['price'])
                            count = count + 1
                            for seisan_data in seisan_datas:
                                if seisan_data['ym'] == month and data2['emp_id'] == seisan_data['emp_id']:
                                    if data2['min_time'] != 0 and data2['max_time'] != 0:
                                        if seisan_data['seisan_time'] > 0:
                                            sh1.cell(row=15+count,column=1,value=count + 1)
                                            sh1.cell(row=15+count,column=2,value="精算" + "【" + data2['first_name'] + data2['last_name'] +  "】")
                                            sh1.cell(row=15+count,column=5,value=seisan_data['seisan_time'])
                                            sh1.cell(row=15+count,column=6,value="時間")
                                            sh1.cell(row=15+count,column=7,value=round(data2['price']/data2['max_time']))
                                        else:
                                            sh1.cell(row=15+count,column=1,value=count + 1)
                                            sh1.cell(row=15+count,column=2,value="精算" + "【" + data2['first_name'] + data2['last_name'] +  "】")
                                            sh1.cell(row=15+count,column=5,value=seisan_data['seisan_time'])
                                            sh1.cell(row=15+count,column=6,value="時間")
                                            sh1.cell(row=15+count,column=7,value=round(data2['price']/data2['min_time']))
                                        count = count + 1
                                        break
                    
                    if data_exists_flag == True:
                        download_file_list.append(filename)
                        # save the excel with name or you can give specific location of your choice
                        wb.save(os.path.join(base_path, filename))
                    wb.close()

    # 注文書
    if service[1:2] == "1":
        
        order_file_datas = []
        with connection.cursor() as cursor:
            cursor.execute(sql_export, (2, start[:6], end[:6]))
            order_file_datas = dictfetchall(cursor)

        for data in order_file_datas:
            data['start_date'] = data['start_date'].strftime('%Y%m%d')
            if data['end_date'] is not None:
                data['end_date'] = data['end_date'].strftime('%Y%m%d')

        months = months_between(start, end)
        output_max_month = months[-1]
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        if len(order_file_datas) > 0:
            
            for month in months:
                # load workbook if you want to write in existing file else use openpyxl.Workbook()
                wb = openpyxl.load_workbook(os.path.join(base_path, "注文書.xlsx"))
                #get the active sheet
                sh1=wb.active
                sh1=wb['注文書']
                done_customer_list = []
                for data1 in order_file_datas:
                    current_customer_id = data1['customer_id']
                    if current_customer_id in done_customer_list:
                        continue
                    done_customer_list.append(current_customer_id)
                    filename = "{}_{}_注文書.xlsx".format(month, data1['customer_name'])
                    count = 0
                    data_exists_flag = False

                    for data2 in order_file_datas:
                        if current_customer_id == data2['customer_id'] and month >= data2['start_date'][:6] and (data['end_date'] is None or month <= data2['end_date'][:6]):

                            data_exists_flag = True
                            # pass which row and column and value which you want to update
                            sh1.cell(row=2,column=1,value=data['customer_name'])
                            sh1.cell(row=2,column=8,value='JCL-' + month + "-" + str(count + 1).zfill(3))
                            sh1.cell(row=3,column=8,value=month[:2] + "/" + month[2:] + "/01")
                            sh1.cell(row=5,column=1,value="件名：" + data['project_name'])
                            # 支払期限
                            current_date = datetime.strptime(month + '01', '%Y%m%d')
                            next_two_month = current_date + relativedelta(months=2)
                            next_month = next_two_month.replace(day=28) + timedelta(days=4)
                            last_day_of_month = next_month - timedelta(days=next_month.day)

                            sh1.cell(row=15+count,column=1,value=count + 1)
                            sh1.cell(row=15+count,column=2,value=data['project_name'] + "【" + data['first_name'] + data['last_name'] +  "】" + month )
                            sh1.cell(row=15+count,column=4,value=1)
                            sh1.cell(row=15+count,column=6,value="人・月")
                            sh1.cell(row=15+count,column=7,value=data['price'])
                            count = count + 1
                    
                    if data_exists_flag == True:
                        download_file_list.append(filename)
                        # save the excel with name or you can give specific location of your choice
                        wb.save(os.path.join(base_path, filename))
                    wb.close()

    with zipfile.ZipFile(os.path.join(base_path,zip_tempfile_path), 'w', compression=zipfile.ZIP_DEFLATED) as new_zip:
        for f_index in download_file_list:
            new_zip.write(os.path.join(base_path, f_index))

    response = HttpResponse(content_type='application/octet-stream')
    temp_filename = zip_tempfile_path.encode("utf8")
    urlenc_filename = urllib.parse.quote(temp_filename)
    response['Content-Disposition'] = 'attachment;filename="{}";filename*=UTF-8\'\'{}'.format(temp_filename, urlenc_filename)

    with open(os.path.join(base_path, zip_tempfile_path), 'rb') as f_in:
        shutil.copyfileobj(f_in, response)

    for f_index in download_file_list:
        if os.path.exists(os.path.join(base_path, f_index)):
            os.remove(os.path.join(base_path, f_index))
    if os.path.exists(os.path.join(base_path,zip_tempfile_path)):
        os.remove(os.path.join(base_path,zip_tempfile_path))

    return response

